# import_export.py
"""
Import / Export helpers for Game Manager GUI
Handles CSV, Excel, JSON, SQLite, TXT files with URL normalization
Automatically fixes IGDB image URLs (\\images.igdb.com → https://images.igdb.com)
"""

from __future__ import annotations
import os
import json
import csv
import sqlite3
import hashlib
import html
import webbrowser
from datetime import datetime
from pathlib import Path
from typing import Tuple, List, Dict, Optional, Any
import tempfile
import warnings
import re
import config
import base64
from cache_utils import _game_cache_dir_for_game

# Suppress warnings
warnings.filterwarnings('ignore')

# Optional libraries - import if available
pd = None
load_workbook = None
REPORTLAB_AVAILABLE = False
FPDF_AVAILABLE = False

# Try pandas for Excel import
try:
    import pandas as pd
except ImportError:
    pd = None

# Try openpyxl for Excel import
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# Try reportlab for PDF export
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import mm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Try fpdf for PDF export (fallback)
try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

# Import sanitize helper from your project
try:
    from utils_sanitize import sanitize_original_title
except ImportError:
    def sanitize_original_title(s: str) -> dict:
        return {"base_title": s, "version": "", "repack": "", "notes": "", "modes": []}

# Default cache directory for storing images
DEFAULT_CACHE_BASE = os.path.join(tempfile.gettempdir(), "game_manager_cache")
os.makedirs(DEFAULT_CACHE_BASE, exist_ok=True)

# Color theme for exports
COLOR_THEME = {
    "primary": "#2c3e50",
    "secondary": "#3498db",
    "success": "#27ae60",
    "warning": "#f39c12",
    "danger": "#e74c3c",
    "light": "#ecf0f1",
    "dark": "#34495e",
    "info": "#3498db",
    "background": "#f8f9fa",
    "border": "#dee2e6"
}

# ----------------------------------------------------------------------
# Helper: empty game template
# ----------------------------------------------------------------------
def empty_game(title: str = "") -> Dict[str, Any]:
    return {
        "title": title,
        "app_id": "",
        "igdb_id": "",
        "release_date": "",
        "developer": "",
        "publisher": "",
        "genres": "",
        "description": "",
        "cover_url": "",
        "trailer_webm": "",
        "screenshots": [],
        "image_cache_paths": [],
        "microtrailer_cache_path": [],
        "shortcut_links": "",
        "steam_link": "",
        "steamdb_link": "",
        "pcgw_link": "",
        "igdb_link": "",
        "save_location": "",
        "savegame_location": [],
        "game_drive": "",
        "scene_repack": "",
        "game_modes": "",
        "original_title": "",
        "original_title_base": "",
        "original_title_version": "",
        "original_notes": "",
        "patch_version": "",
        "player_perspective": "",
        "themes": "",
        "played": False,
        "trailers": [],
    }

# Header mapping for CSV/Excel import
_HEADER_MAP = {
    "title": "title", "steam id": "app_id", "appid": "app_id",
    "release": "release_date", "release date": "release_date",
    "developer": "developer", "publisher": "publisher", "genres": "genres",
    "trailer": "trailer_webm", "screenshots": "screenshots",
    "steam": "steam_link", "steamdb": "steamdb_link", 
    "pcgamingwiki": "pcgw_link", "pcgw": "pcgw_link",
    "igdb": "igdb_link", "igdb id": "igdb_id", "igdb_id": "igdb_id",
    "game drive": "game_drive", "scene/repack": "scene_repack", "scene": "scene_repack",
    "game modes": "game_modes", "modes": "game_modes",
    "original title": "original_title", "original": "original_title",
    "patch/version": "patch_version", "patch": "patch_version", "version": "patch_version",
    "played": "played", "save location": "save_location", "savegame": "save_location",
    "player perspective": "player_perspective", "player_perspective": "player_perspective",
    "themes": "themes", "shortcut_links": "shortcut_links", "shortcut link": "shortcut_links",
    "image_cache_paths": "image_cache_paths", "image_cache_path": "image_cache_paths",
    "savegame_location": "savegame_location", "savegame_locations": "savegame_location",
    "trailers": "trailers", "microtrailers_extra": "microtrailers_extra",
    "microtrailer_cache_path": "microtrailer_cache_path",
}

def normalize_headers(headers: List[str]) -> List[str]:
    return [_HEADER_MAP.get(h.strip().lower(), h.strip().lower()) for h in headers]

def _normalize_url(url: str) -> str:
    if not url or not isinstance(url, str):
        return ""
    url = url.strip().replace("\\", "/")
    if url.startswith("//"):
        url = "https:" + url
    if "images.igdb.com" in url and not url.startswith("http"):
        url = "https://" + url.lstrip("/")
    return url.rstrip("/")

def _enhance_igdb_images(game: Dict) -> Dict:
    cover = game.get("cover_url", "")
    if cover:
        cover = _normalize_url(cover)
        if "images.igdb.com" in cover:
            for size in ["t_thumb", "t_cover_small", "t_cover_big", "t_logo_med", "t_screenshot_med", "t_screenshot_big"]:
                if f"/{size}/" in cover:
                    cover = cover.replace(f"/{size}/", "/t_720p/")
                    break
        game["cover_url"] = cover
    shots = game.get("screenshots", [])
    if shots:
        fixed = []
        for s in shots:
            if s:
                s = _normalize_url(s)
                if "images.igdb.com" in s:
                    for size in ["t_thumb", "t_cover_small", "t_cover_big", "t_logo_med", "t_screenshot_med", "t_screenshot_big"]:
                        if f"/{size}/" in s:
                            s = s.replace(f"/{size}/", "/t_720p/")
                            break
                fixed.append(s)
        game["screenshots"] = fixed
    return game

def _get_thumbnail_data_uri(game: Dict) -> str:
    """Return a data URI for the cover thumbnail, or empty string if not available."""
    # 1. Check cached coverart.jpg using application's cache directory
    cache_dir = _game_cache_dir_for_game(game)
    cover_path = cache_dir / "coverart.jpg"
    if cover_path.exists():
        try:
            with open(cover_path, "rb") as f:
                img_data = base64.b64encode(f.read()).decode('utf-8')
            return f"data:image/jpeg;base64,{img_data}"
        except:
            pass
    
    # 2. Try remote cover_url
    cover_url = game.get("cover_url", "")
    if cover_url:
        return cover_url  # external URL – will need internet
    
    # 3. Placeholder (empty grey square)
    return "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='32' height='32' viewBox='0 0 32 32'%3E%3Crect width='32' height='32' fill='%23cccccc'/%3E%3C/svg%3E"

# ----------------------------------------------------------------------
# Import functions (unchanged except for IGDB ID handling)
# ----------------------------------------------------------------------
def import_excel(path: str) -> Tuple[List[Dict], Optional[str]]:
    try:
        if pd is not None:
            df = pd.read_excel(path, dtype=str).fillna("")
            raw_headers = list(df.columns)
            headers = normalize_headers(raw_headers)
            new_rows = []
            for _, row in df.iterrows():
                game = empty_game()
                for h_raw, h in zip(raw_headers, headers):
                    val = row[h_raw]
                    if pd.isna(val) or val is None or str(val).strip() == "":
                        continue
                    val_str = str(val).strip()
                    if h == "screenshots":
                        game[h] = [s.strip() for s in val_str.split("|") if s.strip()]
                    elif h == "savegame_location":
                        game["savegame_location"] = [s.strip() for s in val_str.split("|") if s.strip()]
                    elif h == "played":
                        game[h] = val_str.lower() in ("yes", "y", "true", "1", "checked")
                    elif h == "image_cache_paths":
                        game[h] = [s.strip() for s in val_str.split("|") if s.strip()]
                    else:
                        game[h] = val_str
                if not game.get("title") and game.get("original_title"):
                    san = sanitize_original_title(game["original_title"])
                    game["title"] = san.get("base_title") or game["original_title"]
                game = _enhance_igdb_images(game)
                new_rows.append(game)
            return new_rows, None
        elif load_workbook is not None:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.rows)
            if not rows:
                return [], "Excel file empty"
            raw_headers = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
            headers = normalize_headers(raw_headers)
            new_rows = []
            for r in rows[1:]:
                game = empty_game()
                for i, cell in enumerate(r):
                    if i >= len(headers):
                        break
                    h = headers[i]
                    val = "" if cell.value is None else str(cell.value).strip()
                    if not val:
                        continue
                    if h == "screenshots":
                        game[h] = [s.strip() for s in val.split("|") if s.strip()]
                    elif h == "savegame_location":
                        game["savegame_location"] = [s.strip() for s in val.split("|") if s.strip()]
                    elif h == "played":
                        game[h] = val.lower() in ("yes", "y", "true", "1", "checked")
                    elif h == "image_cache_paths":
                        game[h] = [s.strip() for s in val.split("|") if s.strip()]
                    else:
                        game[h] = val
                if not game.get("title") and game.get("original_title"):
                    san = sanitize_original_title(game["original_title"])
                    game["title"] = san.get("base_title") or game["original_title"]
                game = _enhance_igdb_images(game)
                new_rows.append(game)
            return new_rows, None
        else:
            return [], "No Excel reader available (install pandas or openpyxl)"
    except Exception as e:
        return [], str(e)

def import_csv(path: str) -> Tuple[List[Dict], Optional[str]]:
    try:
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            raw_headers = next(reader, [])
            headers = normalize_headers(raw_headers)
            new_rows = []
            for row in reader:
                if not row:
                    continue
                game = empty_game()
                for h, val in zip(headers, row):
                    if not val:
                        continue
                    val_str = val.strip()
                    if h == "screenshots":
                        game[h] = [s.strip() for s in val_str.split("|") if s.strip()]
                    elif h == "savegame_location":
                        game["savegame_location"] = [s.strip() for s in val_str.split("|") if s.strip()]
                    elif h == "played":
                        game[h] = val_str.lower() in ("yes", "y", "true", "1", "checked")
                    elif h == "image_cache_paths":
                        game[h] = [s.strip() for s in val_str.split("|") if s.strip()]
                    else:
                        game[h] = val_str
                if not game.get("title") and game.get("original_title"):
                    san = sanitize_original_title(game["original_title"])
                    game["title"] = san.get("base_title") or game["original_title"]
                game = _enhance_igdb_images(game)
                new_rows.append(game)
        return new_rows, None
    except Exception as e:
        return [], str(e)

def import_txt(path: str) -> Tuple[List[Dict], Optional[str]]:
    try:
        new_rows = []
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                original = line.strip()
                if not original:
                    continue
                game = empty_game()
                game["original_title"] = original
                san = sanitize_original_title(original)
                game["original_title_base"] = san.get("base_title", "")
                game["original_title_version"] = san.get("version", "")
                game["scene_repack"] = san.get("repack", "")
                game["original_notes"] = san.get("notes", "")
                game["game_modes"] = ", ".join(san.get("modes", []))
                game["title"] = san.get("base_title") or original
                game = _enhance_igdb_images(game)
                new_rows.append(game)
        return new_rows, None
    except Exception as e:
        return [], str(e)

# ----------------------------------------------------------------------
# JSON functions
# ----------------------------------------------------------------------
def save_to_json(path: str, games: List[Dict]) -> Optional[str]:
    try:
        parent = os.path.dirname(path)
        if parent:
            os.makedirs(parent, exist_ok=True)
        out = []
        for g in games:
            gg = dict(g)
            gg = _enhance_igdb_images(gg)
            gg["screenshots"] = list(gg.get("screenshots") or [])
            gg["image_cache_paths"] = list(gg.get("image_cache_paths") or [])
            gg["savegame_location"] = list(gg.get("savegame_location") or [])
            out.append(gg)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(out, f, indent=2, ensure_ascii=False)
        return None
    except Exception as e:
        return str(e)

def load_from_json(path: str) -> Tuple[List[Dict], Optional[str]]:
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, list):
            return [], "Invalid JSON DB format (expected list)"
        out = []
        for r in data:
            if not isinstance(r, dict):
                continue
            g = empty_game()
            g.update(r)
            g["screenshots"] = list(g.get("screenshots") or [])
            g["image_cache_paths"] = list(g.get("image_cache_paths") or [])
            if isinstance(g.get("savegame_location"), str):
                g["savegame_location"] = [s.strip() for s in g["savegame_location"].split("|") if s.strip()]
            else:
                g["savegame_location"] = list(g.get("savegame_location") or [])
            g = _enhance_igdb_images(g)
            out.append(g)
        return out, None
    except Exception as e:
        return [], str(e)

# ----------------------------------------------------------------------
# SQLite functions
# ----------------------------------------------------------------------
def save_to_sqlite(db_path: str, games: List[Dict]) -> Optional[str]:
    try:
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS games (
                id TEXT PRIMARY KEY,
                title TEXT,
                app_id TEXT,
                igdb_id TEXT,
                release_date TEXT,
                developer TEXT,
                publisher TEXT,
                genres TEXT,
                description TEXT,
                cover_url TEXT,
                trailer_webm TEXT,
                screenshots TEXT,
                image_cache_paths TEXT,
                shortcut_links TEXT,
                steam_link TEXT,
                steamdb_link TEXT,
                pcgw_link TEXT,
                igdb_link TEXT,
                save_location TEXT,
                savegame_location TEXT,
                game_drive TEXT,
                scene_repack TEXT,
                game_modes TEXT,
                original_title TEXT,
                original_title_base TEXT,
                original_title_version TEXT,
                original_notes TEXT,
                patch_version TEXT,
                player_perspective TEXT,
                themes TEXT,
                played INTEGER
            )
        """)
        for g in games:
            g = _enhance_igdb_images(g)
            app_id = str(g.get("app_id", "")).strip()
            igdb_id = str(g.get("igdb_id", "")).strip()
            gid = app_id if app_id and app_id != "Not Found" else (igdb_id if igdb_id else None)
            if not gid:
                gid = hashlib.sha256(json.dumps(g, sort_keys=True).encode("utf-8")).hexdigest()
            shots_json = json.dumps(g.get("screenshots", []), ensure_ascii=False)
            cache_json = json.dumps(g.get("image_cache_paths", []), ensure_ascii=False)
            save_json = json.dumps(g.get("savegame_location", []), ensure_ascii=False)
            played_int = 1 if g.get("played", False) else 0
            c.execute("""
                INSERT OR REPLACE INTO games VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                )
            """, (
                gid, g.get("title", ""), app_id, igdb_id, g.get("release_date", ""),
                g.get("developer", ""), g.get("publisher", ""), g.get("genres", ""),
                g.get("description", ""), g.get("cover_url", ""), g.get("trailer_webm", ""),
                shots_json, cache_json, g.get("shortcut_links", ""), g.get("steam_link", ""),
                g.get("steamdb_link", ""), g.get("pcgw_link", ""), g.get("igdb_link", ""),
                g.get("save_location", ""), save_json, g.get("game_drive", ""),
                g.get("scene_repack", ""), g.get("game_modes", ""), g.get("original_title", ""),
                g.get("original_title_base", ""), g.get("original_title_version", ""),
                g.get("original_notes", ""), g.get("patch_version", ""),
                g.get("player_perspective", ""), g.get("themes", ""), played_int
            ))
        conn.commit()
        conn.close()
        return None
    except Exception as e:
        return str(e)

def load_from_sqlite(db_path: str) -> Tuple[List[Dict], Optional[str]]:
    try:
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute("""
            SELECT title, app_id, igdb_id, release_date, developer, publisher, genres, description,
                   cover_url, trailer_webm, screenshots, image_cache_paths, shortcut_links,
                   steam_link, steamdb_link, pcgw_link, igdb_link, save_location, savegame_location,
                   game_drive, scene_repack, game_modes, original_title, original_title_base,
                   original_title_version, original_notes, patch_version, player_perspective, themes,
                   played
            FROM games
        """)
        rows = c.fetchall()
        conn.close()
        games = []
        for r in rows:
            try:
                shots = json.loads(r[10] or "[]")
            except:
                shots = []
            try:
                cache = json.loads(r[11] or "[]")
            except:
                cache = []
            try:
                save = json.loads(r[18] or "[]")
            except:
                save = []
            game = {
                "title": r[0] or "", "app_id": r[1] or "", "igdb_id": r[2] or "",
                "release_date": r[3] or "", "developer": r[4] or "", "publisher": r[5] or "",
                "genres": r[6] or "", "description": r[7] or "", "cover_url": r[8] or "",
                "trailer_webm": r[9] or "", "screenshots": shots, "image_cache_paths": cache,
                "shortcut_links": r[12] or "", "steam_link": r[13] or "", "steamdb_link": r[14] or "",
                "pcgw_link": r[15] or "", "igdb_link": r[16] or "", "save_location": r[17] or "",
                "savegame_location": save, "game_drive": r[19] or "", "scene_repack": r[20] or "",
                "game_modes": r[21] or "", "original_title": r[22] or "",
                "original_title_base": r[23] or "", "original_title_version": r[24] or "",
                "original_notes": r[25] or "", "patch_version": r[26] or "",
                "player_perspective": r[27] or "", "themes": r[28] or "", "played": bool(r[29])
            }
            game = _enhance_igdb_images(game)
            games.append(game)
        return games, None
    except Exception as e:
        return [], str(e)

# ----------------------------------------------------------------------
# Cache functions
# ----------------------------------------------------------------------
def game_cache_dir(game: Dict, cache_base: Optional[str] = None) -> str:
    base = cache_base or DEFAULT_CACHE_BASE
    basep = Path(base)
    appid = str(game.get("app_id") or "").strip()
    if appid:
        sub = f"game_{appid}"
    else:
        key = (game.get("title") or "") + "|" + (game.get("original_title") or "")
        h = hashlib.sha256(key.encode("utf-8")).hexdigest()[:12]
        sub = f"game_{h}"
    d = basep / sub
    d.mkdir(parents=True, exist_ok=True)
    return str(d)

def prune_game_cache_dir(game: Dict, keep: int = 8, cache_base: Optional[str] = None) -> None:
    d = game_cache_dir(game, cache_base=cache_base)
    p = Path(d)
    if not p.exists():
        return
    files = sorted([f for f in p.iterdir() if f.is_file()], key=lambda f: f.stat().st_mtime, reverse=True)
    for old in files[keep:]:
        try:
            old.unlink()
        except:
            pass

def save_image_bytes(game: Dict, url: str, data: bytes, cache_base: Optional[str] = None) -> Optional[str]:
    try:
        d = game_cache_dir(game, cache_base=cache_base)
        h = hashlib.sha256(url.encode("utf-8")).hexdigest()
        fname = f"{h}.bin"
        path = Path(d) / fname
        with open(path, "wb") as fh:
            fh.write(data)
        prune_game_cache_dir(game, keep=8, cache_base=cache_base)
        return str(path)
    except:
        return None

# ----------------------------------------------------------------------
# Merge function
# ----------------------------------------------------------------------
def merge_imported_rows(existing_games: List[Dict], imported_rows: List[Dict], prefer_imported: bool = True) -> List[Dict]:
    by_app = {}
    by_title = {}
    for g in existing_games:
        aid = str(g.get("app_id") or "").strip()
        if aid:
            by_app[aid] = g
        t = g.get("title")
        if t:
            by_title[t] = g
    for imp in imported_rows:
        imp = _enhance_igdb_images(imp)
        aid = str(imp.get("app_id") or "").strip()
        matched = None
        if aid and aid in by_app:
            matched = by_app[aid]
        else:
            t = imp.get("title")
            if t and t in by_title:
                matched = by_title[t]
        if matched:
            for k, v in imp.items():
                if v is None or v == "":
                    continue
                if prefer_imported:
                    matched[k] = v
                else:
                    if not matched.get(k):
                        matched[k] = v
        else:
            existing_games.append(imp)
    return existing_games


# ----------------------------------------------------------------------
# Export functions (PDF and HTML) with dynamic columns
# ----------------------------------------------------------------------
# ----------------------------------------------------------------------
# Export functions (PDF and HTML) with dynamic columns
# ----------------------------------------------------------------------
def _get_pdf_page_size():
    import configparser
    from config import CONFIG_FILE
    from reportlab.lib.pagesizes import A3, A4, letter, landscape
    cfg = configparser.ConfigParser()
    cfg.read(CONFIG_FILE, encoding='utf-8')
    size_str = cfg.get("Export", "pdf_page_size", fallback="A3 Landscape")
    if size_str == "A3 Landscape":
        return landscape(A3)
    elif size_str == "A4 Landscape":
        return landscape(A4)
    else:
        return landscape(letter)


def export_games_to_pdf(path: str, games: List[Dict], title: Optional[str] = None,
                        description_lines: int = 4,
                        columns: Optional[List[Dict]] = None) -> Optional[str]:
    if not REPORTLAB_AVAILABLE:
        return "ReportLab is not installed. Install with: pip install reportlab"
    return _export_games_to_pdf_reportlab(path, games, title, description_lines, columns)

def _export_games_to_pdf_reportlab(path: str, games: List[Dict], title: Optional[str],
                                   description_lines: int, columns: Optional[List[Dict]]) -> Optional[str]:
    try:
        from reportlab.lib import colors
        
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import mm
        import time
        import html

        if not columns:
            columns = [
                {"key": "title", "header": "Title", "width": 20},
                {"key": "app_id", "header": "Steam ID", "width": 8},
                {"key": "description", "header": "Description", "width": 40},
            ]

        total_pct = sum(col["width"] for col in columns)
        page_size = _get_pdf_page_size()  # A3 landscape
        doc = SimpleDocTemplate(path, pagesize=page_size, leftMargin=8*mm, rightMargin=8*mm,
                                topMargin=12*mm, bottomMargin=12*mm, title=title or "Game Collection Report")
        usable_width = page_size[0] - 16*mm
        col_widths_pts = [usable_width * (col["width"] / total_pct) for col in columns]

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TitleStyle', parent=styles['Title'], fontSize=12,
                                     textColor=colors.HexColor(COLOR_THEME["primary"]), alignment=1, spaceAfter=3)
        subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'], fontSize=8,
                                        textColor=colors.HexColor("#666666"), alignment=1, spaceAfter=8)
        header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontSize=7,
                                      textColor=colors.black, fontName='Helvetica-Bold', alignment=1, leading=8)
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=6, leading=7, wordWrap='CJK')
        small_cell_style = ParagraphStyle('SmallCellStyle', parent=styles['Normal'], fontSize=5, leading=6, wordWrap='CJK')

        total = len(games)
        played = sum(1 for g in games if g.get("played"))
        remaining = total - played
        success_rate = (played / total * 100) if total > 0 else 0

        story = []
        timestamp = time.strftime("%Y-%m-%d at %H:%M:%S")
        story.append(Paragraph("Game Collection Report", title_style))
        story.append(Paragraph(f"Generated on {timestamp}", subtitle_style))

        stats_data = [
            [Paragraph("Total Games", header_style), Paragraph("Played Games", header_style), Paragraph("Remaining Games", header_style)],
            [Paragraph(str(total), cell_style), Paragraph(str(played), cell_style), Paragraph(str(remaining), cell_style)],
            [Paragraph("In collection", cell_style), Paragraph(f"{success_rate:.1f}% completion", cell_style), Paragraph("Games to play", cell_style)]
        ]
        stats_table = Table(stats_data, colWidths=[usable_width/3]*3)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor(COLOR_THEME["primary"])),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTSIZE', (0,0), (-1,0), 8), ('FONTSIZE', (0,1), (-1,1), 10), ('FONTSIZE', (0,2), (-1,2), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#dee2e6")),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 8))

        # Build headers, inserting SN column at front
        headers = [Paragraph(col["header"], header_style) for col in columns]
        headers.insert(0, Paragraph("SN", header_style))
        table_data = [headers]

        desc_idx = next((i for i, col in enumerate(columns) if col["key"] == "description"), None)
        link_opts = _get_link_options()

        # --- duplicate detection for exported games ---
        title_counts = {}
        for g in games:
            t = (g.get("title") or "").strip().lower()
            if t:
                title_counts[t] = title_counts.get(t, 0) + 1
        dup_titles = {t for t, cnt in title_counts.items() if cnt > 1}

        row_colors = []
        for game in games:
            # priority: fav > duplicate > played > unplayed
            if game.get("fav", False):
                raw_color = config.FAVORITE_COLOR
            elif (game.get("title") or "").strip().lower() in dup_titles:
                raw_color = config.DUPLICATE_COLOR
            elif game.get("played", False):
                raw_color = config.PLAYED_COLOR
            else:
                raw_color = config.UNPLAYED_COLOR
            desat = config.desaturate_color(raw_color, config.HIGHLIGHT_DESATURATE_PERCENT)
            row_colors.append(colors.HexColor(desat))
                
        for idx, game in enumerate(games, 1):
            row = []
            for col_idx, col in enumerate(columns):
                key = col["key"]
                if key == "resources":
                    screenshots = list(game.get("screenshots", []) or []) + list(game.get("image_cache_paths", []) or [])
                    trailers = [game.get("trailer_webm")] if game.get("trailer_webm") else []
                    ss_links = [f'<a href="{html.escape(u)}">[{i}]</a>' for i, u in enumerate(screenshots[:4], 1) if u]
                    tt_links = [f'<a href="{html.escape(u)}">[{i}]</a>' for i, u in enumerate(trailers[:2], 1) if u]
                    resources_text = ""
                    if ss_links:
                        resources_text += f'<b>SS:</b> {" ".join(ss_links)}<br/>'
                    if tt_links:
                        resources_text += f'<b>TT:</b> {" ".join(tt_links)}'
                    if not resources_text:
                        resources_text = "None"
                    row.append(Paragraph(resources_text, small_cell_style))
                    continue


                if key == "links":
                    links_html = []
                    if link_opts["steam"]:
                        steam_link = game.get("steam_link", "")
                        if not steam_link and game.get("app_id"):
                            steam_link = f"https://store.steampowered.com/app/{game['app_id']}"
                        if steam_link:
                            links_html.append(f'<a href="{steam_link}">Steam</a>')
                    if link_opts["igdb"]:
                        igdb_link = game.get("igdb_link", "")
                        if not igdb_link and game.get("igdb_id"):
                            igdb_link = f"https://www.igdb.com/games/{game['igdb_id']}"
                        if igdb_link:
                            links_html.append(f'<a href="{igdb_link}">IGDB</a>')
                    if link_opts["pcgw"] and game.get("pcgw_link"):
                        links_html.append(f'<a href="{game["pcgw_link"]}">PCGW</a>')
                    if link_opts["steamdb"] and game.get("steamdb_link"):
                        links_html.append(f'<a href="{game["steamdb_link"]}">SteamDB</a>')
                    links_text = " | ".join(links_html) if links_html else "None"
                    row.append(Paragraph(links_text, cell_style))
                    continue

                
                # Inside the column loop, after handling 'resources' and 'links'...
                value = None
                skip_escape = False

                if key == "patch_version":
                    value = game.get("patch_version", "") or game.get("original_title_version", "")
                elif key == "title":
                    title_text = game.get("title", "")
                    # Add version if available
                    version = game.get("patch_version", "") or game.get("original_title_version", "")
                    if version:
                        title_text = f"{title_text} ({version})"
                    if game.get("played"):
                        title_text = f"{title_text} ✅"
                    if game.get("fav"):
                        title_text = f"{title_text} ♥"
                    # Add user rating (star + rating)
                    rating = game.get("user_rating")
                    if rating is not None and rating != "":
                        try:
                            rating_num = float(rating)
                            title_text = f"{title_text} ★{rating_num:.1f}"
                        except:
                            pass
                    escaped_title = html.escape(title_text)
                    cover_url = game.get("cover_url", "")
                    if cover_url:
                        value = f'<a href="{cover_url}">{escaped_title}</a>'
                    else:
                        value = escaped_title
                    skip_escape = True
                else:
                    value = game.get(key, "")

                # Convert lists, booleans, etc. (but not for already-HTML columns)
                if not skip_escape:
                    if key == "user_rating" and value:
                        try:
                            value = f"{float(value):.1f}"
                        except:
                            pass
                    if key == "description" and description_lines > 0 and desc_idx is not None:
                        desc_width_pts = col_widths_pts[desc_idx]
                        avg_char_width = 5
                        chars_per_line = max(25, int(desc_width_pts / avg_char_width))
                        max_desc_chars = chars_per_line * min(description_lines, 3)
                        value = str(value)[:max_desc_chars] + ("..." if len(str(value)) > max_desc_chars else "")
                    elif isinstance(value, list):
                        value = ", ".join(str(v) for v in value if v)
                    elif isinstance(value, bool):
                        value = "Yes" if value else "No"
                    # Escape plain text
                    value = html.escape(str(value))

                # Now create hyperlinks for app_id and igdb_id if they are plain text (not already linked)
                if not skip_escape and key == "app_id" and value.isdigit():
                    steam_link = game.get("steam_link", "")
                    if not steam_link:
                        steam_link = f"https://store.steampowered.com/app/{value}"
                    value = f'<a href="{steam_link}">{value}</a>'
                elif not skip_escape and key == "igdb_id" and value:
                    igdb_link = game.get("igdb_link", "")
                    if not igdb_link:
                        igdb_link = f"https://www.igdb.com/games/{value}"
                    value = f'<a href="{igdb_link}">{value}</a>'

                row.append(Paragraph(value, cell_style))
            # Prepend row number
            row.insert(0, Paragraph(str(idx), cell_style))
            table_data.append(row)

        sn_width = usable_width * 0.03
        final_widths = [sn_width] + col_widths_pts
        table = Table(table_data, colWidths=final_widths, repeatRows=1)

        # Build style commands: base style + row backgrounds
        style_cmds = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f2f2f2")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.black),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 7),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#dee2e6")),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
        ]

        # Add row background colors (skip header row)
        for i, color in enumerate(row_colors):
            style_cmds.append(('BACKGROUND', (0, i+1), (-1, i+1), color))

        table.setStyle(TableStyle(style_cmds))
        
        story.append(table)
        story.append(Spacer(1, 6))

        footer = Paragraph(f'<font size="7" color="#666666">Report generated by Game Manager • {timestamp} • Total: {total} games • Played: {played} • Remaining: {remaining}</font>', subtitle_style)
        story.append(footer)
        doc.build(story)
        return None
    except Exception as e:
        print(f"PDF export error: {e}")
        import traceback
        traceback.print_exc()
        return f"ReportLab PDF export error: {str(e)}"
        
        
def export_games_to_html(path: str, games: List[Dict], title: Optional[str] = None,
                         open_after: bool = False, description_lines: int = 4,
                         columns: Optional[List[Dict]] = None) -> Optional[str]:
    """Export games to HTML – fixed column widths, no squashing on mobile."""
    try:
        total = len(games)
        played = sum(1 for g in games if g.get("played"))
        remaining = total - played
        success_rate = (played / total * 100) if total > 0 else 0
        timestamp = datetime.now().strftime("%Y-%m-%d at %H:%M:%S")
        link_opts = _get_link_options()

        if not columns:
            columns = [
                {"key": "title", "header": "Title", "width": 10},
                {"key": "app_id", "header": "Steam ID", "width": 5},
                {"key": "igdb_id", "header": "IGDB ID", "width": 5},
                {"key": "patch_version", "header": "Ver", "width": 4},
                {"key": "release_date", "header": "Rel Date", "width": 4},
                {"key": "description", "header": "Description", "width": 25},
                {"key": "game_modes", "header": "Modes", "width": 6},
                {"key": "genres", "header": "Genres", "width": 8},
                {"key": "themes", "header": "Themes", "width": 5},
                {"key": "user_rating", "header": "Rating", "width": 3},
                {"key": "player_perspective", "header": "Perspective", "width": 5},
                {"key": "developer", "header": "Developer", "width": 5},
                {"key": "publisher", "header": "Publisher", "width": 5},
                {"key": "game_drive", "header": "Drive", "width": 5},
                {"key": "scene_repack", "header": "Scene/Repack", "width": 5},
                {"key": "original_title", "header": "Original Title", "width": 10},
                {"key": "resources", "header": "Resources", "width": 10},
                {"key": "links", "header": "External Links", "width": 5},
                {"key": "savegame_location", "header": "Savegame Locations", "width": 5},
            ]

        total_user_pct = sum(col["width"] for col in columns)
        norm_factors = [(col["width"] / total_user_pct) * 0.97 for col in columns]
        sn_width = 0.03

        # --- duplicate detection for exported games ---
        title_counts = {}
        for g in games:
            t = (g.get("title") or "").strip().lower()
            if t:
                title_counts[t] = title_counts.get(t, 0) + 1
        dup_titles = {t for t, cnt in title_counts.items() if cnt > 1}

        rows = []
        for idx, game in enumerate(games, 1):
            # priority: fav > duplicate > played > unplayed
            if game.get("fav", False):
                raw_color = config.FAVORITE_COLOR
            elif (game.get("title") or "").strip().lower() in dup_titles:
                raw_color = config.DUPLICATE_COLOR
            elif game.get("played", False):
                raw_color = config.PLAYED_COLOR
            else:
                raw_color = config.UNPLAYED_COLOR
            bg_color = config.desaturate_color(raw_color, config.HIGHLIGHT_DESATURATE_PERCENT)

            row_cells = []
             # (keep the rest of the loop exactly as before)
            for col in columns:
                key = col["key"]
                cell_content = ""

                if key == "resources":
                    screenshots = list(game.get("screenshots", []) or []) + list(game.get("image_cache_paths", []) or [])
                    trailers = [game.get("trailer_webm")] if game.get("trailer_webm") else []
                    ss_links = [f'<a href="{html.escape(u)}" target="_blank">[{i}]</a>' for i, u in enumerate(screenshots[:5], 1) if u]
                    tt_links = [f'<a href="{html.escape(u)}" target="_blank">[{i}]</a>' for i, u in enumerate(trailers[:3], 1) if u]
                    html_str = ""
                    if ss_links:
                        html_str += f'<div><strong>SS:</strong> {" ".join(ss_links)}</div>'
                    if tt_links:
                        html_str += f'<div><strong>TT:</strong> {" ".join(tt_links)}</div>'
                    if not html_str:
                        html_str = '<span style="color:#f39c12;font-style:italic;">None</span>'
                    cell_content = html_str

                elif key == "links":
                    links_html = []
                    if link_opts["steam"]:
                        steam_link = game.get("steam_link", "")
                        if not steam_link and game.get("app_id"):
                            steam_link = f"https://store.steampowered.com/app/{game['app_id']}"
                        if steam_link:
                            links_html.append(f'<a href="{html.escape(steam_link)}" target="_blank">Steam</a>')
                    if link_opts["igdb"]:
                        igdb_link = game.get("igdb_link", "")
                        if not igdb_link and game.get("igdb_id"):
                            igdb_link = f"https://www.igdb.com/games/{game['igdb_id']}"
                        if igdb_link:
                            links_html.append(f'<a href="{html.escape(igdb_link)}" target="_blank">IGDB</a>')
                    if link_opts["pcgw"] and game.get("pcgw_link"):
                        links_html.append(f'<a href="{html.escape(game["pcgw_link"])}" target="_blank">PCGW</a>')
                    if link_opts["steamdb"] and game.get("steamdb_link"):
                        links_html.append(f'<a href="{html.escape(game["steamdb_link"])}" target="_blank">SteamDB</a>')
                    cell_content = " | ".join(links_html) if links_html else "None"

                elif key == "title":
                    title_text = game.get("title", "")
                    # Add version
                    version = game.get("patch_version", "") or game.get("original_title_version", "")
                    if version:
                        title_text = f"{title_text} ({version})"
                    if game.get("played"):
                        title_text = f"{title_text} ✅"
                    if game.get("fav"):
                        title_text = f"{title_text} ♥"
                    rating = game.get("user_rating")
                    if rating is not None and rating != "":
                        try:
                            rating_num = float(rating)
                            title_text = f"{title_text} ★{rating_num:.1f}"
                        except:
                            pass

                    # Thumbnail handling
                    img_tag = ""
                    if config.EXPORT_THUMBNAILS:
                        thumb_uri = _get_thumbnail_data_uri(game)
                        w = config.EXPORT_THUMBNAIL_WIDTH
                        h = config.EXPORT_THUMBNAIL_HEIGHT
                        img_tag = f'<img src="{thumb_uri}" width="{w}" height="{h}" style="vertical-align: middle; margin-right: 6px;" loading="lazy">'

                    cover_url = game.get("cover_url", "")
                    if cover_url:
                        cell_content = f'<a href="{html.escape(cover_url)}" target="_blank">{img_tag}{html.escape(title_text)}</a>'
                    else:
                        cell_content = f'{img_tag}{html.escape(title_text)}'

                elif key == "patch_version":
                    value = game.get("patch_version", "") or game.get("original_title_version", "")
                    cell_content = html.escape(str(value))

                else:
                    value = game.get(key, "")
                    if key == "user_rating" and value:
                        try:
                            value = f"{float(value):.1f}"
                        except:
                            pass
                    if key == "description":
                        max_desc_chars = 80 * description_lines
                        value = str(value)[:max_desc_chars] + ("..." if len(str(value)) > max_desc_chars else "")
                    if isinstance(value, list):
                        value = ", ".join(str(v) for v in value if v)
                    value = html.escape(str(value))

                    if key == "app_id" and value.isdigit():
                        steam_link = game.get("steam_link", "")
                        if not steam_link:
                            steam_link = f"https://store.steampowered.com/app/{value}"
                        cell_content = f'<a href="{html.escape(steam_link)}" target="_blank">{value}</a>'
                    elif key == "igdb_id" and value:
                        igdb_link = game.get("igdb_link", "")
                        if not igdb_link:
                            igdb_link = f"https://www.igdb.com/games/{value}"
                        cell_content = f'<a href="{html.escape(igdb_link)}" target="_blank">{value}</a>'
                    else:
                        cell_content = value

                row_cells.append(f'<td style="padding:6px; border-bottom:1px solid #dee2e6;">{cell_content}</td>')

            rows.append(f'<tr style="background-color: {bg_color};"><td>{idx}</td>{''.join(row_cells)}</tr>')

        header_cells = ['<th style="width: 3%;">SN</th>']
        for idx, col in enumerate(columns):
            pct = norm_factors[idx] * 100
            header_cells.append(f'<th style="width: {pct:.2f}%;">{html.escape(col["header"])}</th>')
        header_row = ''.join(header_cells)

        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{html.escape(title or "Game Manager Export")}</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{
            margin: 0;
            padding: 20px;
            background: #f0f2f5;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
        }}
        .export-container {{
            max-width: 100%;
            background: white;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            overflow-x: auto;
        }}
        table {{
            width: 100%;
            table-layout: fixed;
            border-collapse: collapse;
            font-size: 13px;
            min-width: 1200px;
        }}
        th, td {{
            padding: 10px 8px;
            border-bottom: 1px solid #e9ecef;
            vertical-align: top;
            word-break: break-word;
        }}
        th {{
            background-color: #f2f4f8;
            font-weight: 600;
            border-bottom: 2px solid #dee2e6;
            text-align: left;
        }}
        a {{
            color: #3498db;
            text-decoration: none;
        }}
        a:hover {{
            text-decoration: underline;
        }}
        .header {{
            background: linear-gradient(to right, #2c3e50, #4a6491);
            color: white;
            padding: 20px;
            border-radius: 12px 12px 0 0;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 16px;
            margin: 20px;
        }}
        .stat-card {{
            background: #f8f9fa;
            padding: 16px;
            border-radius: 8px;
            border-left: 4px solid #3498db;
        }}
        .stat-card.success {{ border-left-color: #27ae60; }}
        .stat-card.warning {{ border-left-color: #f39c12; }}
        .stat-value {{
            font-size: 28px;
            font-weight: bold;
            margin: 8px 0 4px;
        }}
        .footer {{
            margin: 20px;
            padding: 12px;
            background: #f8f9fa;
            border-radius: 8px;
            text-align: center;
            font-size: 12px;
            color: #6c757d;
        }}
        @media (max-width: 768px) {{
            body {{ padding: 10px; }}
            th, td {{ font-size: 11px; padding: 6px; }}
        }}
        @media print {{
            .export-container {{ overflow: visible; }}
            table {{ min-width: auto; }}
        }}
    </style>
</head>
<body>
<div class="export-container">
    <div class="header">
        <h1 style="margin:0; font-size:22px;">Game Collection Report</h1>
        <p style="margin:5px 0 0; opacity:0.9;">Generated on {timestamp}</p>
    </div>
    <div class="stats-grid">
        <div class="stat-card success"><div class="stat-value">{total}</div><p>Total Games</p></div>
        <div class="stat-card"><div class="stat-value">{played}</div><p>Played Games<br>({success_rate:.1f}% completion)</p></div>
        <div class="stat-card warning"><div class="stat-value">{remaining}</div><p>Remaining Games</p></div>
    </div>
    <div style="margin:20px; overflow-x:auto;">
        <table>
            <thead>
                <tr>
                    {header_row}
                </tr>
            </thead>
            <tbody>
                {''.join(rows)}
            </tbody>
        </table>
    </div>
    <div class="footer">
        Report generated by Game Manager • {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} •
        Total: {total} • Played: {played} • Remaining: {remaining}
    </div>
</div>
<script>
    (function() {{
        var btn = document.createElement('button');
        btn.textContent = '🖨️ Print';
        btn.style.cssText = 'position:fixed;bottom:20px;right:20px;padding:10px 16px;background:#2c3e50;color:white;border:none;border-radius:40px;cursor:pointer;font-size:14px;z-index:1000;';
        btn.onclick = () => window.print();
        document.body.appendChild(btn);
    }})();
</script>
</body>
</html>"""
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(html_content, encoding='utf-8')
        if open_after:
            webbrowser.open(p.as_uri())
        return None
    except Exception as e:
        return f"Error exporting to HTML: {str(e)}"
        

def export_file_by_extension(path: str, games: List[Dict], **kwargs) -> Optional[str]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return export_games_to_csv(path, games)
    elif ext in (".xlsx", ".xls"):
        return export_games_to_excel(path, games)
    elif ext == ".pdf":
        return export_games_to_pdf(path, games, kwargs.get('title'), kwargs.get('description_lines', 4), kwargs.get('columns'))
    elif ext in (".html", ".htm"):
        return export_games_to_html(path, games, kwargs.get('title'), kwargs.get('open_after', False),
                                    kwargs.get('description_lines', 4), kwargs.get('columns'))
    elif ext == ".json":
        return save_to_json(path, games)
    elif ext in (".db", ".sqlite"):
        return save_to_sqlite(path, games)
    else:
        return f"Unsupported file extension: {ext}"

def _get_link_options():
    import configparser
    from config import CONFIG_FILE
    cfg = configparser.ConfigParser()
    cfg.read(CONFIG_FILE, encoding='utf-8')
    return {
        "steam": cfg.getboolean("ExportColumns", "links_steam", fallback=True),
        "igdb": cfg.getboolean("ExportColumns", "links_igdb", fallback=True),
        "pcgw": cfg.getboolean("ExportColumns", "links_pcgw", fallback=True),
        "steamdb": cfg.getboolean("ExportColumns", "links_steamdb", fallback=True),
    }


# ----------------------------------------------------------------------
# Convenience function
# ----------------------------------------------------------------------
def import_file_by_extension(path: str) -> Tuple[List[Dict], Optional[str]]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return import_csv(path)
    elif ext in (".xlsx", ".xls"):
        return import_excel(path)
    elif ext in (".txt", ".list"):
        return import_txt(path)
    elif ext == ".json":
        return load_from_json(path)
    elif ext in (".db", ".sqlite"):
        return load_from_sqlite(path)
    else:
        return [], f"Unsupported extension: {ext}"

     
# CSV and Excel export (unchanged but added igdb_id)
def export_games_to_csv(path: str, games: List[Dict]) -> Optional[str]:
    try:
        fields = ["title", "app_id", "igdb_id", "release_date", "developer", "publisher", "genres", "description",
                  "cover_url", "trailer_webm", "steam_link", "steamdb_link", "pcgw_link", "igdb_link",
                  "game_drive", "scene_repack", "game_modes", "original_title", "original_title_base",
                  "original_title_version", "original_notes", "patch_version", "player_perspective",
                  "themes", "played", "save_location", "shortcut_links"]
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            for game in games:
                row = {}
                for field in fields:
                    value = game.get(field)
                    if isinstance(value, list):
                        row[field] = "|".join(str(item) for item in value if item)
                    elif isinstance(value, bool):
                        row[field] = "Yes" if value else "No"
                    else:
                        row[field] = str(value or "")
                writer.writerow(row)
        return None
    except Exception as e:
        return f"Error exporting to CSV: {str(e)}"

def export_games_to_excel(path: str, games: List[Dict]) -> Optional[str]:
    if pd is None:
        return "pandas not available"
    try:
        fields = ["title", "app_id", "igdb_id", "release_date", "developer", "publisher", "genres", "description",
                  "cover_url", "trailer_webm", "steam_link", "steamdb_link", "pcgw_link", "igdb_link",
                  "game_drive", "scene_repack", "game_modes", "original_title", "original_title_base",
                  "original_title_version", "original_notes", "patch_version", "player_perspective",
                  "themes", "played", "save_location", "shortcut_links", "screenshots", "image_cache_paths", "savegame_location"]
        data = []
        for game in games:
            row = {}
            for field in fields:
                value = game.get(field)
                if isinstance(value, list):
                    row[field] = "|".join(str(item) for item in value if item)
                elif isinstance(value, bool):
                    row[field] = "Yes" if value else "No"
                else:
                    row[field] = str(value or "")
            data.append(row)
        df = pd.DataFrame(data, columns=fields)
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Games', index=False)
            worksheet = writer.sheets['Games']
            for column in worksheet.columns:
                max_len = 0
                col_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except:
                        pass
                adjusted = min(max_len + 2, 50)
                worksheet.column_dimensions[col_letter].width = adjusted
        return None
    except Exception as e:
        return f"Error exporting to Excel: {str(e)}"

# ----------------------------------------------------------------------
# CLI test (unchanged)
# ----------------------------------------------------------------------
def test_cli():
    import sys
    if len(sys.argv) < 2:
        print("Usage: python import_export.py <command> [options]")
        return
    command = sys.argv[1].lower()
    if command == "import" and len(sys.argv) >= 3:
        path = sys.argv[2]
        games, err = import_file_by_extension(path)
        if err:
            print(f"Error: {err}")
        else:
            print(f"Imported {len(games)} games")
    # other commands similar...
if __name__ == "__main__":
    test_cli()